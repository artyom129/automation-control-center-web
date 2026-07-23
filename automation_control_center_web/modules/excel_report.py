from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill


def create_excel_report(clean_df: pd.DataFrame, invalid_df: pd.DataFrame, output_path: Path, report_title: str) -> None:
    output_path.parent.mkdir(exist_ok=True)

    wb = Workbook()
    dashboard = wb.active
    dashboard.title = "Dashboard"

    clean_sheet = wb.create_sheet("Clean Records")
    invalid_sheet = wb.create_sheet("Invalid Records")

    write_dashboard(dashboard, report_title, clean_df, invalid_df)
    write_dataframe(clean_sheet, clean_df)
    write_dataframe(invalid_sheet, invalid_df)

    wb.save(output_path)


def write_dashboard(ws, report_title: str, clean_df: pd.DataFrame, invalid_df: pd.DataFrame) -> None:
    ws["A1"] = report_title
    ws["A1"].font = Font(size=18, bold=True, color="1F4E78")
    ws.merge_cells("A1:B1")

    metrics = [
        ["Metric", "Value"],
        ["Clean records", len(clean_df)],
        ["Invalid records", len(invalid_df)],
        ["Total records", len(clean_df) + len(invalid_df)],
        ["Columns", len(clean_df.columns) if not clean_df.empty else 0],
    ]

    for row in metrics:
        ws.append(row)

    style_header(ws, 2)
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 16

    chart = BarChart()
    chart.title = "Automation Summary"
    chart.y_axis.title = "Count"

    data = Reference(ws, min_col=2, min_row=3, max_row=6)
    categories = Reference(ws, min_col=1, min_row=3, max_row=6)
    chart.add_data(data, titles_from_data=False)
    chart.set_categories(categories)
    chart.height = 8
    chart.width = 14

    ws.add_chart(chart, "D2")


def write_dataframe(ws, df: pd.DataFrame) -> None:
    if df.empty:
        ws.append(["No records"])
        return

    ws.append(list(df.columns))

    for row in df.itertuples(index=False):
        ws.append(list(row))

    style_header(ws, 1)
    ws.freeze_panes = "A2"

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:
            max_length = max(max_length, len(str(cell.value or "")))

        ws.column_dimensions[column_letter].width = min(max_length + 3, 36)


def style_header(ws, row_number: int) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)

    for cell in ws[row_number]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
